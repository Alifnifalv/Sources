import docx
import requests
import json
import re
import json5
import os
from dotenv import load_dotenv
import pandas as pd
import io
from openpyxl import load_workbook
import subprocess
import tempfile
import speech_recognition as sr
from word2number import w2n
import fitz

load_dotenv(override=True)
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
MODEL_NAME = os.getenv("MODEL_NAME", "anthropic/claude-3-opus-20240229")
PROMPT_FILE_PATH = os.getenv("PROMPT_FILE_PATH", "curriculum_analysis_prompt.txt")

if not OPENROUTER_API_KEY:
    raise ValueError("OPENROUTER_API_KEY is not set. Please configure your .env file.")

if not MODEL_NAME:
    raise ValueError("MODEL_NAME is not set in .env. Please set a default model.")

# --- Constants ---
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xls', 'xlsx'}

def allowed_file(filename):
    """Check if the file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_prompt_template():
    """Load the prompt template from the external file"""
    try:
        with open(PROMPT_FILE_PATH, 'r', encoding='utf-8') as file:
            return file.read()
    except FileNotFoundError:
        raise FileNotFoundError(f"Prompt file not found: {PROMPT_FILE_PATH}")
    except Exception as e:
        raise Exception(f"Error loading prompt file: {str(e)}")

def create_prompt(document_text, class_labels=None, subject_labels=None):
    """Create the prompt by filling in the template"""
    class_labels = class_labels or ["Class:", "Std.:", "Grade:"]
    subject_labels = subject_labels or ["Subject:", "Course:"]
    
    # Load the template
    template = load_prompt_template()
    
    # Replace placeholders
    prompt = template.replace("{document_text}", document_text)
    prompt = prompt.replace("{class_labels}", ', '.join(class_labels))
    prompt = prompt.replace("{subject_labels}", ', '.join(subject_labels))
    
    return prompt

def generate_response_openrouter(document_text):
    """Generate response using OpenRouter API with robust error handling."""
    try:
        # Generate the prompt using the external template
        prompt = create_prompt(document_text)
        
        response = requests.post(
            url="https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json",
                "HTTP-Referer": "https://your-api-domain.com",  # IMPORTANT: Replace!
                "X-Title": "Document Class Details Extractor API",
            },
            json={
                "model": MODEL_NAME,
                "messages": [
                    {
                        "role": "system",
                        "content": "You are a precise JSON generator. Extract and return ONLY the requested information in valid JSON format, adhering strictly to the provided schema. Do not include any extra text, explanations, or markdown formatting."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                "temperature": 0.2   # Lower temperature for more consistent output
            },
            timeout=300
        )

        if not response.ok:
            return {"error": f"API request failed with status code {response.status_code}", "details": response.text}, response.status_code

        try:
            response_json = response.json()
        except json.JSONDecodeError:
            return {"error": "Failed to decode API response as JSON", "details": response.text}, 500

        if "choices" not in response_json or not response_json["choices"]:
            return {"error": "API response missing expected 'choices' field", "details": response_json}, 500

        content = response_json["choices"][0]["message"]["content"].strip()
        
        # Improved JSON extraction - handles multiple patterns
        json_pattern = re.compile(r'```(?:json)?(.*?)```|({[\s\S]*})', re.DOTALL)
        matches = json_pattern.findall(content)
        
        for match in matches:
            for submatch in match:
                if submatch.strip():
                    cleaned_json = submatch.strip()
                    try:
                        parsed_json = json.loads(cleaned_json)
                        return parsed_json, 200
                    except json.JSONDecodeError:
                        continue
        
        # If we haven't returned yet, try to find any JSON-like structure
        try:
            # Remove all non-JSON content before the first '{' and after the last '}'
            first_brace = content.find('{')
            last_brace = content.rfind('}')
            if first_brace != -1 and last_brace != -1:
                json_str = content[first_brace:last_brace+1]
                parsed_json = json5.loads(json_str)
                return parsed_json, 200
        except Exception:
            pass
            
        # If all else fails
        return {
            "error": "Failed to extract valid JSON from response",
            "details": content
        }, 400

    except requests.RequestException as e:
        return {"error": f"Request failed: {e}"}, 500

def extract_text_from_pdf(pdf_file_path_or_stream):
    try:
        doc = fitz.open(stream=pdf_file_path_or_stream.read(), filetype="pdf") # Open from stream
        # OR doc = fitz.open(pdf_file_path) # If you have a path
        all_text = []
        for page in doc:
            all_text.append(page.get_text())
        doc.close()
        extracted_text = "\n".join(all_text)

        if not extracted_text.strip():
            return None, "No extractable text found in the PDF (PyMuPDF)."
        return extracted_text, None
    except Exception as e:
         return None, str(e)

def extract_text_from_excel(excel_file):
    """Extracts text from all sheets of an Excel file, preserving merged cell values."""
    try:
        # Load the workbook
        excel_data = io.BytesIO(excel_file.read())
        wb = load_workbook(excel_data, data_only=True)

        all_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Store merged cell values
            merged_cell_values = {}
            for merged_range in ws.merged_cells.ranges:
                first_cell = ws[merged_range.start_cell.coordinate]
                merged_value = first_cell.value

                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_cell_values[(row, col)] = merged_value

            # Convert worksheet to DataFrame
            data = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                new_row = []
                for col_idx, cell in enumerate(row, start=1):
                    # Fill merged cell values if needed
                    value = merged_cell_values.get((row_idx, col_idx), cell)
                    new_row.append(str(value) if value is not None else "")  # Convert None to empty string
                data.append(new_row)

            df = pd.DataFrame(data)
            
            for row in df.values.tolist():
                all_text.append("\t".join(row))  # Ensures all elements are strings

        return "\n".join(all_text), None
    except Exception as e:
        return None, str(e)

def get_document_text(files):
    """Extract text from uploaded files (PDF, DOCX, and Excel)."""
    all_text = []

    for file_key in files:
        file = files[file_key]
        file_text = None
        error = None

        if file.filename.lower().endswith(".pdf"):
            file_text, error = extract_text_from_pdf(file)
        elif file.filename.lower().endswith(".docx"):
            try:
                doc = docx.Document(file)
                file_text = "\n".join([p.text for p in doc.paragraphs])
            except Exception as e:
                error = str(e)
        elif file.filename.lower().endswith((".xls", ".xlsx")):
            file_text, error = extract_text_from_excel(file)

        if error:
            return None, error
        if file_text:
            all_text.append(file_text)

    if not all_text:
        return None, "No text extracted from any files."
    return "\n\n".join(all_text), None
    
def convert_webm_to_wav(input_file, wav_path):
    """Convert WebM audio to WAV format using FFmpeg."""
    try:
        subprocess.run([
            "ffmpeg", "-y", "-i", input_file,
            "-ar", "16000", "-ac", "1", "-af", "afftdn",  # Fix here
            "-f", "wav", wav_path
        ], check=True)
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"FFmpeg conversion failed: {e}")

def transcribe_audio(file_obj):
    """Convert WebM to WAV, then transcribe."""
    # Create a temporary file to save the uploaded file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".webm") as temp_webm:
        # If it's a BytesIO object, write its contents
        if hasattr(file_obj, 'read'):
            temp_webm.write(file_obj.read())
        else:
            # If it's a file path, just use it directly
            temp_webm_path = file_obj
        
        temp_webm_path = temp_webm.name

    # Create a temporary WAV file 
    wav_path = temp_webm_path.replace(".webm", ".wav")
    
    try:
        # Convert WebM to WAV
        convert_webm_to_wav(temp_webm_path, wav_path)
        
        # Initialize recognizer
        recognizer = sr.Recognizer()
        with sr.AudioFile(wav_path) as source:
            audio = recognizer.record(source)  # Read entire file
        
        try:
            # Use Google's free Speech-to-Text API
            text = recognizer.recognize_google(audio)
            #text = ' '.join(str(w2n.word_to_num(word)) if word.isdigit() == False else word for word in text.split())
            words = text.split()
            contains_number = any(word.isdigit() or word.lower() in w2n.american_number_system for word in words)

            if contains_number:
                text = ' '.join(str(w2n.word_to_num(word)) if word.lower() in w2n.american_number_system else word for word in words)
            
            return text
        except sr.UnknownValueError:
            return "Speech recognition could not understand the audio."
        except sr.RequestError:
            return "Could not request results from Google Speech Recognition service."
    
    finally:
        # Clean up temporary files
        try:
            os.remove(temp_webm_path)
            os.remove(wav_path)
        except Exception:
            pass


def interact_with_rasa(user_input):
    """Send user input to Rasa and get the response."""
    rasa_url = "http://localhost:5005/webhooks/rest/webhook"
    payload = {"sender": "user", "message": user_input}
    try:
        response = requests.post(rasa_url, json=payload)
        if response.status_code == 200:
            return response.json()
        return "I'm sorry, I didn't understand that."
    except Exception as e:
        return f"Error connecting to Rasa: {e}"
